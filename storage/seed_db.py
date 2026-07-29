import os
import re
import json
import zipfile
import random
from openpyxl import load_workbook

# Paths
downloads_dir = r"C:\Users\user\Downloads"
excel_file = next((f for f in os.listdir(downloads_dir) if f.endswith("20260719.xlsx")), None)
if not excel_file:
    raise FileNotFoundError("Could not find any file ending with 20260719.xlsx in Downloads")
EXCEL_PATH = os.path.join(downloads_dir, excel_file)
ZIP_PATH = os.path.join(downloads_dir, "Image.zip")
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
STATIC_IMAGE_DIR = os.path.join(PROJECT_ROOT, "backend", "backend", "app", "static", "images")
MOCK_DB_PATH = os.path.join(PROJECT_ROOT, "storage", "mock_db.json")

def parse_dong(address):
    """
    Extracts the local 'dong/eup/myeon' administrative district from the address string.
    e.g., '세종특별자치시 전의면 비암사길 137' -> '전의면'
    """
    if not address:
        return "기타"
    
    tokens = address.split()
    # Search for token ending with eup, myeon, or dong
    for token in tokens:
        if token.endswith(('읍', '면', '동')):
            return token
            
    # Fallback to second token if first is city
    if len(tokens) >= 2:
        return tokens[1]
        
    return "기타"

def normalize_era(era_str):
    """
    Normalizes the era text to a standard list for graph and charts mapping.
    """
    if not era_str:
        return "조선시대"
        
    era_str = str(era_str).strip()
    
    if "삼국" in era_str:
        return "삼국시대"
    if "신라" in era_str:
        return "통일신라시대"
    if "고려" in era_str:
        return "고려시대"
    if "조선" in era_str:
        return "조선시대"
    if "근대" in era_str or "일제" in era_str or "대한제국" in era_str or "19" in era_str:
        return "근대"
    if "현대" in era_str or "20" in era_str:
        return "현대"
        
    return era_str

def infer_category(name, desc):
    """
    Infers the heritage category based on keywords in name or description.
    """
    full_text = (name + " " + desc).lower()
    
    if "무형" in full_text:
        return "무형문화재"
    if "은행나무" in full_text or "나무" in full_text or "숲" in full_text or "천연기념물" in full_text:
        return "기념물"
    if "정자" in full_text or "서원" in full_text or "가옥" in full_text:
        return "문화재자료"
    if "공원" in full_text or "호수" in full_text or "휴양림" in full_text or "체험" in full_text:
        return "현대명소"
        
    return "유형문화재"

def run_seeding():
    print("=== SEJONG DATA SEEDING ENGINE START ===")
    
    # 1. Create target directories if they don't exist
    os.makedirs(STATIC_IMAGE_DIR, exist_ok=True)
    
    # 2. Extract Heritage Data from Excel
    print(f"Reading Excel: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH, read_only=True)
    sheet = wb.active
    
    excel_records = []
    # Identify headers in row 1
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    print(f"Excel Headers found: {headers}")
    
    # Map headers to indices
    header_to_idx = {h: idx for idx, h in enumerate(headers)}
    
    # Read rows
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]: # Skip empty rows
            continue
            
        h_id = str(row[header_to_idx["H_ID"]]).strip()
        name = str(row[header_to_idx["명칭"]]).strip()
        address = str(row[header_to_idx["소재지"]]).strip()
        desc = str(row[header_to_idx["소개"]]).strip() if row[header_to_idx["소개"]] else ""
        era_raw = str(row[header_to_idx["시대"]]).strip() if row[header_to_idx["시대"]] else ""
        thinking_text = str(row[header_to_idx["생각할 거리"]]).strip() if row[header_to_idx["생각할 거리"]] else ""
        
        # Parse dong and normalize era
        dong = parse_dong(address)
        era = normalize_era(era_raw)
        category = infer_category(name, desc)
        
        # Geolocation Simulation (Sejong City bounds: Lat 36.4~36.6, Lng 127.15~127.35)
        # Create deterministic coordinates based on H_ID to maintain repeatability
        random.seed(h_id)
        latitude = round(random.uniform(36.45, 36.60), 6)
        longitude = round(random.uniform(127.18, 127.32), 6)
        
        excel_records.append({
            "id": h_id,
            "name": name,
            "category": category,
            "era": era,
            "address": address,
            "dong": dong,
            "latitude": latitude,
            "longitude": longitude,
            "description": desc,
            "thought_prompt": thinking_text,
            "image_url": f"/static/images/{h_id}.jpg",
            "views": random.randint(10, 150)
        })
        
    print(f"Extracted {len(excel_records)} records from Excel.")
    assert len(excel_records) == 119, f"Expected 119 excel records, but found {len(excel_records)}"
    
    # 3. Unzip and Verify Images
    print(f"Reading Images ZIP: {ZIP_PATH}")
    extracted_images_count = 0
    
    with zipfile.ZipFile(ZIP_PATH, 'r') as zip_ref:
        namelist = zip_ref.namelist()
        
        # Filter image files
        image_files = [
            f for f in namelist 
            if f.lower().endswith(('.jpg', '.jpeg')) and not f.startswith('__MACOSX')
        ]
        print(f"Found {len(image_files)} image files in ZIP.")
        
        # Extract and map matching images
        for f in image_files:
            filename = os.path.basename(f)
            # Match H1.jpg to H119.jpg
            match = re.match(r"^(H\d+)\.(jpg|jpeg)$", filename, re.IGNORECASE)
            if match:
                h_id = match.group(1).upper()
                target_filename = f"{h_id}.jpg"
                target_path = os.path.join(STATIC_IMAGE_DIR, target_filename)
                
                # Write flat file structure
                with open(target_path, "wb") as out_f:
                    out_f.write(zip_ref.read(f))
                extracted_images_count += 1
                
    print(f"Extracted and saved {extracted_images_count} images to: {STATIC_IMAGE_DIR}")
    assert extracted_images_count == 119, f"Expected 119 extracted images, but got {extracted_images_count}"
    
    # 4. Generate Graph Connections for local routing simulation
    print("Generating simulated graph connections...")
    graph_connections = []
    # Build local routes connecting heritages in the same dong/district
    dong_map = {}
    for h in excel_records:
        dong_map.setdefault(h["dong"], []).append(h["id"])
        
    random.seed("sejong_graph")
    for dong, ids in dong_map.items():
        if len(ids) >= 2:
            # Connect adjacent nodes sequentially
            for i in range(len(ids) - 1):
                graph_connections.append({
                    "source_id": ids[i],
                    "target_id": ids[i+1],
                    "distance_km": round(random.uniform(0.8, 3.5), 1),
                    "travel_time_mins": random.randint(3, 12)
                })
                
    # 5. Populate Citizen candidates, reviews and course structures in mock_db.json
    mock_db = {
        "cultural_heritages": excel_records,
        "citizen_reports": [
            {
                "id": 1,
                "reporter_name": "김세종",
                "title": "금강 보람교 교각 야경",
                "description": "금강 보람교의 현대적 아치교 디자인과 LED 컬러 야경이 조화를 이루는 명소입니다.",
                "address": "세종특별자치시 보람동 금강변",
                "latitude": 36.4862,
                "longitude": 127.2915,
                "historical_significance": "신도시의 대표적 교량 조형미 및 여가 유산",
                "status": "PENDING",
                "admin_comment": None,
                "created_at": "2026-07-23T10:00:00Z",
                "updated_at": "2026-07-23T10:00:00Z"
            }
        ],
        "citizen_heritage_candidates": [
            {
                "id": 1,
                "name": "조치원 정수장 문화정원",
                "image_urls": ["/static/images/candidates/c1_1.jpg"],
                "latitude": 36.6025,
                "longitude": 127.2981,
                "description": "옛 정수시설을 조경 정원과 전시실로 탈바꿈한 근대 건축 재생 명소입니다.",
                "reporter_id": "user_demo_1",
                "votes": 12,
                "status": "PENDING",
                "admin_comment": None,
                "created_at": "2026-07-23T11:00:00Z"
            }
        ],
        "heritage_reviews": [
            {
                "id": 1,
                "heritage_id": "H1",
                "user_id": "user_demo_2",
                "image_url": "/static/images/reviews/r1.jpg",
                "content": "비암사의 아늑함과 조용한 산사 분위기가 아주 좋습니다. 사계절 추천해요!",
                "created_at": "2026-07-23T12:00:00Z"
            }
        ],
        "user_courses": [
            {
                "id": "c8932912-9d48-4a50-a466-ca22c2bffac3",
                "user_id": "user_demo_1",
                "title": "전의면 전통 역사 걷기",
                "description": "전의면 근방의 고즈넉한 비암사와 전적지들을 둘러보는 도보 힐링 코스입니다.",
                "theme": "역사 탐방",
                "transit_type": "보행",
                "duration_mins": 90,
                "generated_content": "동화 형식: 옛날 전의면 비암사 깊은 숲속에는...",
                "stops": [
                    { "stop_order": 1, "heritage_id": "H1" }
                ],
                "created_at": "2026-07-23T13:00:00Z",
                "updated_at": "2026-07-23T13:00:00Z"
            }
        ],
        "user_recommendation_status": [
            {
                "candidate_id": 1,
                "user_id": "user_demo_1",
                "recommended_date": "2026-07-23T11:00:00Z",
                "status": "PENDING",
                "feedback": "심사 대기 상태입니다."
            }
        ],
        "graph_connections": graph_connections
    }
    
    # Save mock database
    print(f"Saving Seeded Database to: {MOCK_DB_PATH}")
    with open(MOCK_DB_PATH, "w", encoding="utf-8") as f:
        json.dump(mock_db, f, ensure_ascii=False, indent=2)
        
    print("=== SEJONG DATA SEEDING ENGINE COMPLETED (SUCCESS) ===")

if __name__ == "__main__":
    run_seeding()
